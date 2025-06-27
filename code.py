import pandas as pd
import os
from dateutil import parser
import matplotlib.pyplot as plt
from openpyxl import load_workbook

if os.path.exists('messy_file.csv'):
    df = pd.read_csv('messy_file.csv')
    df = df.apply(lambda col: col.str.strip() if col.dtype == "object" else col)
    df['Amount'] = df["Amount"].replace(r'(?i)(rs|inr|₹|usd|\$)[\.\s]*', '', regex=True)
    df['Amount'] = df["Amount"].replace(r'[^\d\-\.]', '', regex=True)
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
    df['Category'] = df['Category'].astype(str).str.capitalize()
    df['Category'] = df['Category'].replace('Transpot', 'Transport')

    def smart_parse_date(date_str):
        try:
            return parser.parse(date_str, dayfirst=True)
        except:
            return pd.NaT

    df['Date'] = df['Date'].apply(smart_parse_date)
    df['Date'] = df['Date'].dt.date
    df = df.dropna(subset=['Date', 'Amount', 'Category'])

    cat_totals = df.groupby('Category')['Amount'].sum()
    total = cat_totals.sum()
    top_cats = cat_totals.sort_values(ascending=False).head(3)
    
    print("Here's Your Cleaned CSV File:")
    print(df)

    
     
    cat_totals.plot(kind='bar', color='orange')
    plt.title("Expenses by Category")
    plt.ylabel("Amount (₹)")
    plt.xlabel("Category")
    plt.tight_layout()
    for i, val in enumerate(cat_totals):
        plt.text(i, val, f'₹{val:.0f}', ha='center', va='bottom', fontsize=8)
    plt.savefig('chart_exp.png')
    plt.show()

    with pd.ExcelWriter('cleaned_sheet.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cleaned Data', index=False)

    wb = load_workbook('cleaned_sheet.xlsx')
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
    wb.save('cleaned_sheet.xlsx')

    with open("summary.txt", "w", encoding="utf-8") as f:
        f.write("Expense Summary Report\n")
        f.write("=========================\n")
        f.write(f"Total Spend: ₹{total:.2f}\n\n")
        f.write("Top 3 Categories:\n")
        for cat, amt in top_cats.items():
            f.write(f"- {cat}: ₹{amt:.2f}\n")

    

    

else:
    print('File NOT Found')
